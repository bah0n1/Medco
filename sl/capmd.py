import openpyxl



def save(counter,medicine_name ="",small_form="",Generic_name="",Strength_name="",Manufactured_name="",unt_price="",pack_price="",Indications="",Therapeutic="",Pharmacology="",Dosage="",Interaction="",Contraindications="",Side_Effects="",Pregnancy="",Precautions="",Populations="",Overdose="",Storage=""):
    wb = openpyxl.load_workbook("medicine1.xlsx")
    sheet= wb["Sheet"]
    sheet.cell(counter,1,medicine_name)
    sheet.cell(counter,2,small_form)
    sheet.cell(counter,3,Generic_name)
    sheet.cell(counter,4,Strength_name)
    sheet.cell(counter,5,Manufactured_name)
    sheet.cell(counter,6,unt_price)
    sheet.cell(counter,7,pack_price)
    sheet.cell(counter,8,Indications)
    sheet.cell(counter,9,Therapeutic)
    sheet.cell(counter,10,Pharmacology)
    sheet.cell(counter,11,Dosage)
    sheet.cell(counter,12,Interaction)
    sheet.cell(counter,13,Contraindications)
    sheet.cell(counter,14,Side_Effects)
    sheet.cell(counter,15,Pregnancy)
    sheet.cell(counter,16,Precautions)
    sheet.cell(counter,17,Populations)
    sheet.cell(counter,18,Overdose)
    sheet.cell(counter,19,Storage)
    wb.save("medicine1.xlsx")


