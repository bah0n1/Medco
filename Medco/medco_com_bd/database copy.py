import openpyxl

class Database:
    def __init__(self,counter):
        self.counter=counter

    def print_data_test(self):
        wb=openpyxl.load_workbook("medicine.xlsx")
        sheet= wb.active
        if self.counter <= sheet.max_row:
            medicine_name=sheet.cell(self.counter,1).value
            small_form=sheet.cell(self.counter,2).value
            Generic_name=sheet.cell(self.counter,3).value
            Strength_name=sheet.cell(self.counter,4).value
            Manufactured_name=sheet.cell(self.counter,5).value
            unt_price=sheet.cell(self.counter,6).value
            pack_price=sheet.cell(self.counter,7).value
            Indications=sheet.cell(self.counter,8).value
            Therapeutic=sheet.cell(self.counter,9).value
            Pharmacology=sheet.cell(self.counter,10).value
            Dosage=sheet.cell(self.counter,11).value
            Interaction=sheet.cell(self.counter,12).value
            Contraindications=sheet.cell(self.counter,13).value
            Side_Effects=sheet.cell(self.counter,14).value
            Pregnancy=sheet.cell(self.counter,15).value
            Precautions=sheet.cell(self.counter,16).value
            Populations=sheet.cell(self.counter,17).value
            Overdose=sheet.cell(self.counter,18).value
            Storage=sheet.cell(self.counter,19).value
            lku=sheet.cell(self.counter,20).value
            
            
            return(medicine_name,small_form,Generic_name,Strength_name,Manufactured_name,
            unt_price,pack_price,Indications,Therapeutic,Pharmacology,Dosage,Interaction,Contraindications,Side_Effects,
            Pregnancy,Precautions,Populations,Overdose,Storage,lku)
        else:
            return "There have no more row"

# cc=Database(18133)
# kk=cc.print_data_test()

